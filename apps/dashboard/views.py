from openpyxl.styles import Font, Border, Side
from openpyxl.styles.builtins import output
from rest_framework.response import Response
from rest_framework.generics import ListAPIView
from django.db.models import Sum, ExpressionWrapper, F, FloatField, OuterRef, Subquery
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.views import APIView

from .filters import StockDashboardFilter
from config.pagination import CustomPageNumberPagination
from .serializers import ItemsDashboardSerializer
from ..items.models import Product, Stock, MeasurementProduct
from django.db.models.functions import Cast
from openpyxl import Workbook
from django.http import HttpResponse
from io import BytesIO


class ItemsDashboardAPIView(ListAPIView):
    filter_backends = (DjangoFilterBackend,)
    filterset_class = StockDashboardFilter
    pagination_class = CustomPageNumberPagination

    def get_queryset(self):
        qs = Stock.objects.select_related(
            "product", "store", "supplier"
        ).only(
            'id', 'product',
            'quantity', 'quantity_for_history',
            'store', 'supplier', 'date_of_arrived'
        )
        if self.request.user.is_superuser:
            return qs.all()
        elif self.request.user.role in ['Администратор', 'Продавец']:
            return qs.filter(store=self.request.user.store)
        else:
            return qs.none()

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()

        filtered_queryset = self.filter_queryset(queryset).exclude(product__category__category_name='Рейка')
        total_product = filtered_queryset.count()

        non_reyka_qs = filtered_queryset.values(
            'product__product_name', 'store__name'

        ).annotate(total_quantity=Sum('quantity'), total_kub_volume=Sum("total_volume")).order_by(
            'product__product_name')

        total_volume = filtered_queryset.filter(quantity__gt=0).aggregate(Sum('total_volume'))

        metr_subquery = MeasurementProduct.objects.filter(
            product=OuterRef('product'),
            measurement__measurement_name='Метр'
        ).annotate(
            metr_value=Cast(F('number'), FloatField())
        ).values('metr_value')[:1]

        reyka_qs = self.filter_queryset(queryset).filter(product__category__category_name='Рейка', ).values(
            'product__product_name', 'store__name'
        ).annotate(total_kub=Sum(
            ExpressionWrapper(
                ((F('quantity') / Subquery(metr_subquery, output_field=FloatField())) * F('product__kub')),
                output_field=FloatField()
            )
        ), total_quantity=Sum('quantity'))

        combined_data = list(non_reyka_qs) + list(reyka_qs)
        total_reyka = reyka_qs.count()
        total_volume_reyka = reyka_qs.filter(quantity__gt=0).aggregate(Sum('total_kub'))
        total = (total_volume['total_volume__sum'] or 0) + (total_volume_reyka['total_kub__sum'] or 0)

        page = self.paginate_queryset(combined_data)
        if page is not None:
            return self.get_paginated_response({
                "total_product": total_product,
                "info_products": list(page),
                "total_volume": total_volume['total_volume__sum'] or 0,
                'total_volume_reyka': total_volume_reyka,
                "total": total

            })

        return Response({
            "total_product": total_product,
            "info_products": list(combined_data),
            "reyks": list(reyka_qs),
            "total_volume": total_volume['total_volume__sum'] or 0,
            "total": total
        })


def get_items_dashboard_data(request):
    queryset = Stock.objects.select_related(
        "product", "store", "supplier"
    ).only(
        'id', 'product',
        'quantity', 'quantity_for_history',
        'store', 'supplier', 'date_of_arrived'
    )
    filterset = StockDashboardFilter(request.GET, queryset=queryset)

    if not filterset.is_valid():
        return {
            "info_products": [],
            "total_product": 0,
        }

    filtered_queryset = filterset.qs.exclude(product__category__category_name='Рейка')
    total_product = filtered_queryset.count()

    non_reyka_qs = filtered_queryset.values(
        'product__product_name', 'store__name'

    ).annotate(total_quantity=Sum('quantity'), total_kub_volume=Sum("total_volume")).order_by(
        'product__product_name')

    total_volume = filtered_queryset.filter(quantity__gt=0).aggregate(Sum('total_volume'))

    metr_subquery = MeasurementProduct.objects.filter(
        product=OuterRef('product'),
        measurement__measurement_name='Метр'
    ).annotate(
        metr_value=Cast(F('number'), FloatField())
    ).values('metr_value')[:1]

    reyka_qs = filterset.qs.filter(product__category__category_name='Рейка', ).values(
        'product__product_name', 'store__name'
    ).annotate(total_kub=Sum(
        ExpressionWrapper(
            ((F('quantity') / Subquery(metr_subquery, output_field=FloatField())) * F('product__kub')),
            output_field=FloatField()
        )
    ), total_quantity=Sum('quantity'))

    combined_data = list(non_reyka_qs) + list(reyka_qs)
    total_reyka = reyka_qs.count()
    total_volume_reyka = reyka_qs.filter(quantity__gt=0).aggregate(Sum('total_kub'))
    total = (total_volume['total_volume__sum'] or 0) + (total_volume_reyka['total_kub__sum'] or 0)

    return {
        "total_product": total_product,
        "info_products": list(combined_data),
        "reyks": list(reyka_qs),
        "total_volume": total_volume['total_volume__sum'] or 0,
        "total": total
    }


class ExportExcelAPIView(APIView):
    def get(self, request, *args, **kwargs):
        data = get_items_dashboard_data(request)
        wb = Workbook()
        ws = wb.active
        ws.title = 'Остаток Товаров'
        header_font = Font(bold=True, size=14)
        body_font = Font(size=12)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        headers = ["Название товара", "Магазин", "Количество", "Объём (м3)"]
        ws.append(headers)
        for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(headers)):
            for cell in col:
                cell.font = header_font
                cell.border = thin_border

        for row_data in data['info_products']:
            row = [
                row_data.get("product__product_name"),
                row_data.get("store__name"),
                row_data.get("total_quantity"),
                row_data.get("total_kub_volume") or row_data.get("total_kub") or 0
            ]
            ws.append(row)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4, ):
            for cell in row:
                cell.font = body_font
                cell.border = thin_border

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="dashboard.xlsx"'
        return response
